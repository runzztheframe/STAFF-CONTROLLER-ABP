import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import json
import shutil
from PIL import Image, ImageTk, ImageOps
import os
from datetime import datetime, timedelta
import sys
import subprocess
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
import tempfile
import webbrowser
from tkinter import scrolledtext
import requests
from io import BytesIO
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import seaborn as sns
import calendar
from dateutil.relativedelta import relativedelta

###############################################################################
# CONFIG SYSTEM
###############################################################################

CONFIG_FILE = "config.json"

# MODERN COLOR SCHEME
COLORS = {
    "primary": "#6366f1",
    "primary_dark": "#4f46e5",
    "secondary": "#f43f5e",
    "accent": "#06b6d4",
    "background": "#0f172a",
    "surface": "#1e293b",
    "text_primary": "#f8fafc",
    "text_secondary": "#cbd5e1",
    "success": "#10b981",
    "warning": "#f59e0b",
    "error": "#ef4444"
}

def load_config():
    if not os.path.exists(CONFIG_FILE):
        return None
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

def save_config(data):
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save config: {e}")

def choose_data_folder_dialog(parent=None, title="ជ្រើសរើសថតទិន្នន័យបុគ្គលិក"):
    return filedialog.askdirectory(parent=parent, title=title)

def choose_data_folder_and_save(parent=None):
    folder = choose_data_folder_dialog(parent=parent)
    if folder:
        save_config({"data_folder": folder})
        return folder
    return None

def get_data_folder_from_config():
    config = load_config()
    if config and "data_folder" in config:
        return config["data_folder"]
    return None

###############################################################################
# MODERN LOGIN WINDOW
###############################################################################
class ModernLoginWindow:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("🔐 ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក")
        self.root.geometry("450x600")
        self.root.resizable(False, False)
        self.root.configure(bg=COLORS["background"])
        
        # Center window
        self.root.eval('tk::PlaceWindow . center')
        
        self.setup_ui()
        self.root.mainloop()

    def setup_ui(self):
        # Main container
        main_frame = tk.Frame(self.root, bg=COLORS["background"], padx=40, pady=40)
        main_frame.pack(fill="both", expand=True)
        
        # Logo/Header
        header_frame = tk.Frame(main_frame, bg=COLORS["background"])
        header_frame.pack(pady=(0, 30))
        
        tk.Label(header_frame, text="🏢", font=("Arial", 48), 
                bg=COLORS["background"], fg=COLORS["primary"]).pack()
        
        tk.Label(header_frame, text="ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក", font=("Arial", 24, "bold"), 
                bg=COLORS["background"], fg=COLORS["text_primary"]).pack(pady=(10, 5))
        
        tk.Label(header_frame, text="កំណែ 2.0 - ការគ្រប់គ្រងដ៏ទំនើប", font=("Arial", 12), 
                bg=COLORS["background"], fg=COLORS["text_secondary"]).pack()
        
        # Login Form
        form_frame = tk.Frame(main_frame, bg=COLORS["surface"], padx=30, pady=30, relief="flat", bd=0)
        form_frame.pack(fill="x", pady=20)
        
        # Username
        tk.Label(form_frame, text="👤 ឈ្មោះអ្នកប្រើប្រាស់", font=("Arial", 10, "bold"), 
                bg=COLORS["surface"], fg=COLORS["text_secondary"], anchor="w").pack(fill="x", pady=(0, 5))
        
        self.user_var = tk.StringVar()
        user_entry = tk.Entry(form_frame, textvariable=self.user_var, font=("Arial", 12), 
                             bg=COLORS["background"], fg=COLORS["text_primary"], 
                             insertbackground=COLORS["text_primary"], relief="flat", bd=0)
        user_entry.pack(fill="x", pady=(0, 15), ipady=8)
        user_entry.bind("<FocusIn>", lambda e: user_entry.configure(bg="#374151"))
        user_entry.bind("<FocusOut>", lambda e: user_entry.configure(bg=COLORS["background"]))
        
        # Password
        tk.Label(form_frame, text="🔒 លេខសម្ងាត់", font=("Arial", 10, "bold"), 
                bg=COLORS["surface"], fg=COLORS["text_secondary"], anchor="w").pack(fill="x", pady=(0, 5))
        
        self.pass_var = tk.StringVar()
        pass_entry = tk.Entry(form_frame, textvariable=self.pass_var, font=("Arial", 12), 
                             show="•", bg=COLORS["background"], fg=COLORS["text_primary"],
                             insertbackground=COLORS["text_primary"], relief="flat", bd=0)
        pass_entry.pack(fill="x", pady=(0, 25), ipady=8)
        pass_entry.bind("<FocusIn>", lambda e: pass_entry.configure(bg="#374151"))
        pass_entry.bind("<FocusOut>", lambda e: pass_entry.configure(bg=COLORS["background"]))
        
        # Login Button
        login_btn = tk.Button(form_frame, text="🚀 ចូលប្រើប្រាស់", font=("Arial", 12, "bold"),
                             bg=COLORS["primary"], fg="white", relief="flat", bd=0,
                             command=self.check_login, cursor="hand2")
        login_btn.pack(fill="x", ipady=12, pady=(0, 15))
        login_btn.bind("<Enter>", lambda e: login_btn.configure(bg=COLORS["primary_dark"]))
        login_btn.bind("<Leave>", lambda e: login_btn.configure(bg=COLORS["primary"]))
        
        # Quick Login Hint
        hint_frame = tk.Frame(form_frame, bg=COLORS["surface"])
        hint_frame.pack(fill="x", pady=(10, 0))
        
        tk.Label(hint_frame, text="💡 គ្រាន់តែបំពេញពត៌មានណាមួយរួចចុចចូល", font=("Arial", 9),
                bg=COLORS["surface"], fg=COLORS["accent"]).pack()
        
        # Footer
        footer_frame = tk.Frame(main_frame, bg=COLORS["background"])
        footer_frame.pack(fill="x", pady=(20, 0))
        
        tk.Label(footer_frame, text="🔐 កំណែសាកល្បង • v2.0", font=("Arial", 9),
                bg=COLORS["background"], fg=COLORS["text_secondary"]).pack()

    def check_login(self):
        username = self.user_var.get().strip()
        password = self.pass_var.get().strip()

        # FAKE LOGIN - ACCEPTS ANY INPUT
        if username or password:
            welcome_name = username if username else "អ្នកប្រើប្រាស់"
            messagebox.showinfo("ស្វាគមន៍", f"👋 សូមស្វាគមន៍ {welcome_name}!\n\nបានចូលប្រើប្រាស់ប្រព័ន្ធដោយជោគជ័យ")
            
            self.root.destroy()
            MainApplication()
        else:
            messagebox.showinfo("បញ្ជាក់", "💡 សូមបំពេញពត៌មានមុនពេលចូលប្រើប្រាស់")

###############################################################################
# ENHANCED STAFF MANAGEMENT SYSTEM WITH CALENDAR & CLOCK
###############################################################################
class EnhancedStaffManagementSystem:
    def __init__(self, root, data_folder=None):
        self.root = root
        self.root.title("🏢 ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក v2.0")
        self.root.geometry("1400x900")
        self.root.configure(bg=COLORS["background"])

        # Initialize data folder
        folder = data_folder or get_data_folder_from_config()
        if not folder or not os.path.exists(folder):
            folder = choose_data_folder_and_save(parent=self.root)
            if not folder:
                messagebox.showerror("កំហុស", "ត្រូវការថតទិន្នន័យ!")
                self.root.destroy()
                return

        self.data_folder = folder
        self.photo_folder = os.path.join(self.data_folder, "Photos")
        self.doc_folder = os.path.join(self.data_folder, "Documents")
        self.backup_folder = os.path.join(self.data_folder, "Backups")
        os.makedirs(self.photo_folder, exist_ok=True)
        os.makedirs(self.doc_folder, exist_ok=True)
        os.makedirs(self.backup_folder, exist_ok=True)
        
        self.db_file = os.path.join(self.data_folder, "staff_data.db")

        self.selected_staff_id = None
        self.photo_image = None
        self.current_view = "list"  # list, stats, calendar
        self.current_date = datetime.now()
        
        # Clock variables
        self.time_var = tk.StringVar()
        self.date_var = tk.StringVar()

        self.init_database()
        self.upgrade_database()
        self.create_enhanced_ui()
        self.load_staff_data()
        self.create_backup()
        self.update_clock()  # Start the clock

    def upgrade_database(self):
        """Add missing columns to existing database"""
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # Check if columns exist and add them if they don't
            cursor.execute("PRAGMA table_info(staff)")
            existing_columns = [column[1] for column in cursor.fetchall()]
            
            columns_to_add = [
                ('phone', 'TEXT'),
                ('email', 'TEXT'),
                ('department', 'TEXT'),
                ('created_date', 'TEXT DEFAULT CURRENT_TIMESTAMP'),
                ('last_updated', 'TEXT DEFAULT CURRENT_TIMESTAMP')
            ]
            
            for column_name, column_type in columns_to_add:
                if column_name not in existing_columns:
                    cursor.execute(f"ALTER TABLE staff ADD COLUMN {column_name} {column_type}")
                    print(f"Added column: {column_name}")
            
            conn.commit()
            conn.close()
            
        except Exception as e:
            print(f"Database upgrade error: {e}")

    def init_database(self):
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS staff(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    birthdate TEXT,
                    marital_status TEXT,
                    location TEXT,
                    position TEXT,
                    salary REAL DEFAULT 0,
                    hire_date TEXT,
                    phone TEXT,
                    email TEXT,
                    department TEXT,
                    photo_path TEXT,
                    document_path TEXT,
                    created_date TEXT DEFAULT CURRENT_TIMESTAMP,
                    last_updated TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Create indexes for better performance
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_name ON staff(name)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_department ON staff(department)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_position ON staff(position)")
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            messagebox.showerror("Database Error", f"Cannot initialize database: {e}")
            return False

    def update_clock(self):
        """Update the real-time clock"""
        now = datetime.now()
        current_time = now.strftime("%H:%M:%S")
        current_date = now.strftime("%Y-%m-%d %A")
        
        self.time_var.set(f"⏰ {current_time}")
        self.date_var.set(f"📅 {current_date}")
        
        # Update every second
        self.root.after(1000, self.update_clock)

    def create_enhanced_ui(self):
        """Create enhanced modern UI with clock"""
        # Main container
        main_container = tk.Frame(self.root, bg=COLORS["background"])
        main_container.pack(fill="both", expand=True, padx=15, pady=15)

        # Header with quick stats and clock
        self.create_header(main_container)
        
        # Navigation tabs
        self.create_navigation(main_container)
        
        # Content area
        self.content_frame = tk.Frame(main_container, bg=COLORS["background"])
        self.content_frame.pack(fill="both", expand=True, pady=10)
        
        # Show default view
        self.show_list_view()

    def create_header(self, parent):
        """Create header with quick stats and real-time clock"""
        header_frame = tk.Frame(parent, bg=COLORS["surface"], height=100)
        header_frame.pack(fill="x", pady=(0, 10))
        header_frame.pack_propagate(False)
        
        # Title section
        title_frame = tk.Frame(header_frame, bg=COLORS["surface"])
        title_frame.pack(side="left", fill="both", expand=True, padx=20)
        
        tk.Label(title_frame, text="🏢 ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក", 
                font=("Arial", 20, "bold"), bg=COLORS["surface"], 
                fg=COLORS["text_primary"]).pack(anchor="w")
        
        tk.Label(title_frame, text="កំណែ 2.0 - ការគ្រប់គ្រងដ៏សាមញ្ញ និងមានប្រសិទ្ធភាព", 
                font=("Arial", 10), bg=COLORS["surface"], 
                fg=COLORS["text_secondary"]).pack(anchor="w")
        
        # Right section with clock and stats
        right_frame = tk.Frame(header_frame, bg=COLORS["surface"])
        right_frame.pack(side="right", padx=20)
        
        # Real-time clock
        clock_frame = tk.Frame(right_frame, bg=COLORS["surface"])
        clock_frame.pack(anchor="e")
        
        tk.Label(clock_frame, textvariable=self.time_var, 
                font=("Arial", 12, "bold"), bg=COLORS["surface"], 
                fg=COLORS["accent"]).pack(anchor="e")
        
        tk.Label(clock_frame, textvariable=self.date_var, 
                font=("Arial", 10), bg=COLORS["surface"], 
                fg=COLORS["text_secondary"]).pack(anchor="e")
        
        # Quick stats
        stats_frame = tk.Frame(right_frame, bg=COLORS["surface"])
        stats_frame.pack(anchor="e", pady=(5, 0))
        
        self.total_staff_var = tk.StringVar(value="សរុប: 0")
        self.department_count_var = tk.StringVar(value="នាយកដ្ឋាន: 0")
        
        tk.Label(stats_frame, textvariable=self.total_staff_var, 
                font=("Arial", 10, "bold"), bg=COLORS["surface"], 
                fg=COLORS["success"]).pack(anchor="e")
        
        tk.Label(stats_frame, textvariable=self.department_count_var, 
                font=("Arial", 9), bg=COLORS["surface"], 
                fg=COLORS["accent"]).pack(anchor="e")

    def create_navigation(self, parent):
        """Create navigation tabs"""
        nav_frame = tk.Frame(parent, bg=COLORS["surface"], height=50)
        nav_frame.pack(fill="x", pady=(0, 10))
        nav_frame.pack_propagate(False)
        
        buttons = [
            ("📋 បញ្ជីបុគ្គលិក", self.show_list_view),
            ("📊 របាយការណ៍", self.show_stats_view),
            ("📅 ប្រតិទិន", self.show_calendar_view),
            ("⚙️ ការកំណត់", self.show_settings_view)
        ]
        
        for text, command in buttons:
            btn = tk.Button(nav_frame, text=text, font=("Arial", 10),
                          bg=COLORS["surface"], fg=COLORS["text_primary"],
                          relief="flat", bd=0, command=command)
            btn.pack(side="left", padx=5)
            
            # Highlight current view
            if text.startswith("📋"):
                btn.configure(bg=COLORS["primary"], fg="white")

    def show_calendar_view(self):
        """Show interactive calendar view"""
        self.clear_content()
        self.current_view = "calendar"
        
        calendar_frame = tk.Frame(self.content_frame, bg=COLORS["surface"], padx=20, pady=20)
        calendar_frame.pack(fill="both", expand=True)
        
        # Calendar header
        header_frame = tk.Frame(calendar_frame, bg=COLORS["surface"])
        header_frame.pack(fill="x", pady=(0, 20))
        
        tk.Label(header_frame, text="📅 ប្រតិទិនបុគ្គលិក", 
                font=("Arial", 18, "bold"), bg=COLORS["surface"], 
                fg=COLORS["text_primary"]).pack(side="left")
        
        # Navigation controls
        nav_frame = tk.Frame(header_frame, bg=COLORS["surface"])
        nav_frame.pack(side="right")
        
        tk.Button(nav_frame, text="⬅️ ខែមុន", font=("Arial", 9),
                 bg=COLORS["primary"], fg="white", relief="flat",
                 command=self.previous_month).pack(side="left", padx=5)
        
        self.month_year_var = tk.StringVar()
        month_label = tk.Label(nav_frame, textvariable=self.month_year_var,
                              font=("Arial", 12, "bold"), bg=COLORS["surface"],
                              fg=COLORS["text_primary"])
        month_label.pack(side="left", padx=15)
        
        tk.Button(nav_frame, text="ខែបន្ទាប់ ➡️", font=("Arial", 9),
                 bg=COLORS["primary"], fg="white", relief="flat",
                 command=self.next_month).pack(side="left", padx=5)
        
        tk.Button(nav_frame, text="📅 ថ្ងៃនេះ", font=("Arial", 9),
                 bg=COLORS["accent"], fg="white", relief="flat",
                 command=self.today).pack(side="left", padx=5)
        
        # Create calendar
        self.create_calendar_widget(calendar_frame)
        
        # Staff birthdays section
        self.create_birthdays_section(calendar_frame)
        
        # Update calendar display
        self.update_calendar()

    def create_calendar_widget(self, parent):
        """Create the calendar grid"""
        # Days of week header
        days_frame = tk.Frame(parent, bg=COLORS["surface"])
        days_frame.pack(fill="x", pady=(0, 10))
        
        days = ["អាទិត្យ", "ចន្ទ", "អង្គារ", "ពុធ", "ព្រហស្បតិ៍", "សុក្រ", "សៅរ៍"]
        for day in days:
            tk.Label(days_frame, text=day, font=("Arial", 10, "bold"),
                    bg=COLORS["primary"], fg="white",
                    width=10, height=2).pack(side="left", padx=1, fill="x", expand=True)
        
        # Calendar grid
        self.calendar_frame = tk.Frame(parent, bg=COLORS["surface"])
        self.calendar_frame.pack(fill="both", expand=True)
        
        # Initialize day buttons
        self.day_buttons = []
        for row in range(6):
            row_buttons = []
            for col in range(7):
                btn = tk.Button(self.calendar_frame, text="", font=("Arial", 10),
                              bg=COLORS["surface"], fg=COLORS["text_primary"],
                              relief="flat", width=10, height=3)
                btn.grid(row=row, column=col, padx=1, pady=1, sticky="nsew")
                row_buttons.append(btn)
            self.day_buttons.append(row_buttons)
        
        # Configure grid weights
        for i in range(6):
            self.calendar_frame.rowconfigure(i, weight=1)
        for i in range(7):
            self.calendar_frame.columnconfigure(i, weight=1)

    def create_birthdays_section(self, parent):
        """Create staff birthdays section"""
        birthdays_frame = tk.LabelFrame(parent, text="🎂 ខួបកំណើតខែនេះ",
                                      font=("Arial", 12, "bold"), bg=COLORS["surface"],
                                      fg=COLORS["text_primary"], padx=15, pady=15)
        birthdays_frame.pack(fill="x", pady=(20, 0))
        
        self.birthdays_text = scrolledtext.ScrolledText(birthdays_frame, 
                                                      height=6,
                                                      font=("Arial", 10),
                                                      bg=COLORS["background"],
                                                      fg=COLORS["text_primary"],
                                                      relief="flat")
        self.birthdays_text.pack(fill="both", expand=True)
        self.birthdays_text.config(state=tk.DISABLED)

    def update_calendar(self):
        """Update calendar display for current month"""
        year = self.current_date.year
        month = self.current_date.month
        
        # Update month-year label
        month_names = ["មករា", "កុម្ភៈ", "មីនា", "មេសា", "ឧសភា", "មិថុនា", 
                      "កក្កដា", "សីហា", "កញ្ញា", "តុលា", "វិច្ឆិកា", "ធ្នូ"]
        self.month_year_var.set(f"{month_names[month-1]} {year}")
        
        # Clear all buttons
        for row in self.day_buttons:
            for btn in row:
                btn.config(text="", bg=COLORS["surface"], state="disabled")
        
        # Get calendar data
        cal = calendar.monthcalendar(year, month)
        today = datetime.now().date()
        
        # Fill calendar
        for week_num, week in enumerate(cal):
            for day_num, day in enumerate(week):
                if day != 0:
                    btn = self.day_buttons[week_num][day_num]
                    btn.config(text=str(day), state="normal",
                             command=lambda d=day: self.on_day_click(d))
                    
                    # Highlight today
                    if (year == today.year and month == today.month and day == today.day):
                        btn.config(bg=COLORS["accent"], fg="white")
                    else:
                        btn.config(bg=COLORS["surface"], fg=COLORS["text_primary"])
        
        # Update birthdays
        self.update_birthdays_display()

    def update_birthdays_display(self):
        """Update birthdays display for current month"""
        year = self.current_date.year
        month = self.current_date.month
        
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT name, birthdate, position 
                FROM staff 
                WHERE strftime('%m', birthdate) = ? 
                ORDER BY strftime('%d', birthdate)
            """, (f"{month:02d}",))
            
            birthdays = cursor.fetchall()
            conn.close()
            
            self.birthdays_text.config(state=tk.NORMAL)
            self.birthdays_text.delete(1.0, tk.END)
            
            if birthdays:
                for name, birthdate, position in birthdays:
                    if birthdate:
                        bday = datetime.strptime(birthdate, "%Y-%m-%d")
                        age = year - bday.year
                        self.birthdays_text.insert(tk.END, 
                                                 f"🎂 {name} - {position}\n"
                                                 f"   📅 {birthdate} (អាយុ {age} ឆ្នាំ)\n\n")
            else:
                self.birthdays_text.insert(tk.END, "មិនមានខួបកំណើតក្នុងខែនេះទេ\n")
                
            self.birthdays_text.config(state=tk.DISABLED)
            
        except Exception as e:
            print(f"Error loading birthdays: {e}")

    def previous_month(self):
        """Navigate to previous month"""
        self.current_date = self.current_date - relativedelta(months=1)
        self.update_calendar()

    def next_month(self):
        """Navigate to next month"""
        self.current_date = self.current_date + relativedelta(months=1)
        self.update_calendar()

    def today(self):
        """Navigate to current month"""
        self.current_date = datetime.now()
        self.update_calendar()

    def on_day_click(self, day):
        """Handle day click event"""
        year = self.current_date.year
        month = self.current_date.month
        selected_date = f"{year}-{month:02d}-{day:02d}"
        
        # Show staff with birthdays on this day
        self.show_staff_birthdays(selected_date)

    def show_staff_birthdays(self, date):
        """Show staff who have birthdays on selected date"""
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT name, position, department, birthdate 
                FROM staff 
                WHERE strftime('%m-%d', birthdate) = strftime('%m-%d', ?)
            """, (date,))
            
            staff_list = cursor.fetchall()
            conn.close()
            
            if staff_list:
                message = f"បុគ្គលិកដែលមានខួបកំណើតនៅថ្ងៃ {date}:\n\n"
                for name, position, department, birthdate in staff_list:
                    age = datetime.now().year - datetime.strptime(birthdate, "%Y-%m-%d").year
                    message += f"• {name} - {position} ({department})\n  អាយុ {age} ឆ្នាំ\n\n"
            else:
                message = f"មិនមានបុគ្គលិកណាមានខួបកំណើតនៅថ្ងៃ {date} ទេ"
                
            messagebox.showinfo("ខួបកំណើត", message)
            
        except Exception as e:
            messagebox.showerror("កំហុស", f"មិនអាចផ្ទុកទិន្នន័យ: {e}")

    def create_backup(self):
        """Create automatic backup"""
        try:
            backup_file = os.path.join(self.backup_folder, f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
            shutil.copy2(self.db_file, backup_file)
            
            # Keep only last 5 backups
            backups = sorted([f for f in os.listdir(self.backup_folder) if f.startswith('backup_')])
            if len(backups) > 5:
                for old_backup in backups[:-5]:
                    os.remove(os.path.join(self.backup_folder, old_backup))
        except Exception as e:
            print(f"Backup failed: {e}")

    # ... (Keep all the existing methods for list view, stats view, etc.)

    def show_list_view(self):
        """Show staff list view"""
        self.clear_content()
        self.current_view = "list"
        
        # Content area with form and list
        content = tk.Frame(self.content_frame, bg=COLORS["background"])
        content.pack(fill="both", expand=True)
        
        # Left - Quick form
        left_frame = tk.LabelFrame(content, text="👤 បំពេញពត៌មាន",
                                 font=("Arial", 11, "bold"), bg=COLORS["surface"],
                                 fg=COLORS["text_primary"], padx=15, pady=15)
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        # Right - Staff list
        right_frame = tk.LabelFrame(content, text="📋 បញ្ជីបុគ្គលិក", 
                                  font=("Arial", 11, "bold"), bg=COLORS["surface"],
                                  fg=COLORS["text_primary"], padx=15, pady=15)
        right_frame.pack(side="right", fill="both", expand=True, padx=(10, 0))
        
        self.create_quick_form(left_frame)
        self.create_enhanced_list(right_frame)

    def create_quick_form(self, parent):
        """Create quick entry form"""
        form = tk.Frame(parent, bg=COLORS["surface"])
        form.pack(fill="both", expand=True)

        # Form variables
        self.name_var = tk.StringVar()
        self.birthdate_var = tk.StringVar()
        self.marital_var = tk.StringVar()
        self.location_var = tk.StringVar()
        self.position_var = tk.StringVar()
        self.salary_var = tk.StringVar()
        self.hire_date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.phone_var = tk.StringVar()
        self.email_var = tk.StringVar()
        self.department_var = tk.StringVar()
        self.photo_path_var = tk.StringVar()
        self.document_path_var = tk.StringVar()

        # Create form fields in a grid
        fields = [
            ("ឈ្មោះពេញ *", self.name_var, "entry"),
            ("ថ្ងៃខែឆ្នាំកំណើត", self.birthdate_var, "entry"),
            ("ស្ថានភាពគ្រួសារ", self.marital_var, "combo"),
            ("ទីកន្លែង", self.location_var, "entry"),
            ("តួនាទី", self.position_var, "combo"),
            ("ប្រាក់ខែ ($)", self.salary_var, "entry"),
            ("ថ្ងៃចូលធ្វើការ", self.hire_date_var, "entry"),
            ("លេខទូរស័ព្ទ", self.phone_var, "entry"),
            ("អ៊ីមែល", self.email_var, "entry"),
            ("នាយកដ្ឋាន", self.department_var, "combo")
        ]

        for i, (label, var, field_type) in enumerate(fields):
            row = i // 2
            col = (i % 2) * 2
            
            tk.Label(form, text=label, font=("Arial", 9), 
                    bg=COLORS["surface"], fg=COLORS["text_primary"]).grid(
                    row=row, column=col, sticky="w", pady=5, padx=(0, 10))
            
            if field_type == "combo":
                if label == "ស្ថានភាពគ្រួសារ":
                    values = ["នៅលីវ", "រៀបការ", "ផ្សេងៗ"]
                elif label == "តួនាទី":
                    values = ["អ្នកគ្រប់គ្រង", "អ្នកអភិវឌ្ឍ", "អ្នករចនា", "លក់", "HR", "គាំទ្រ", "អនុវិទ្យ", "ផ្សេងៗ"]
                else:
                    values = ["IT", "លក់", "ទីផ្សារ", "HR", "ហិរញ្ញវត្ថុ", "ផ្សេងៗ"]
                
                combo = ttk.Combobox(form, textvariable=var, values=values, 
                                   state="readonly", width=20, font=("Arial", 9))
                combo.grid(row=row, column=col+1, sticky="w", pady=5)
            else:
                entry = tk.Entry(form, textvariable=var, font=("Arial", 9), 
                               width=22, bg="#374151", fg=COLORS["text_primary"],
                               insertbackground=COLORS["text_primary"], relief="flat")
                entry.grid(row=row, column=col+1, sticky="w", pady=5)

        # Photo and document section
        photo_row = (len(fields) + 1) // 2
        tk.Label(form, text="រូបភាព", font=("Arial", 9), 
                bg=COLORS["surface"], fg=COLORS["text_primary"]).grid(
                row=photo_row, column=0, sticky="w", pady=5)
        
        photo_btn_frame = tk.Frame(form, bg=COLORS["surface"])
        photo_btn_frame.grid(row=photo_row, column=1, columnspan=3, sticky="w", pady=5)
        
        tk.Button(photo_btn_frame, text="📁 ជ្រើសរូប", font=("Arial", 8),
                 bg=COLORS["primary"], fg="white", relief="flat",
                 command=self.browse_photo).pack(side="left", padx=2)
        
        tk.Button(photo_btn_frame, text="👀 មើល", font=("Arial", 8),
                 bg=COLORS["accent"], fg="white", relief="flat",
                 command=self.preview_current_photo).pack(side="left", padx=2)

        # Action buttons
        action_row = photo_row + 1
        action_frame = tk.Frame(form, bg=COLORS["surface"])
        action_frame.grid(row=action_row, column=0, columnspan=4, pady=15)
        
        actions = [
            ("➕ បន្ថែម", COLORS["success"], self.add_staff),
            ("✏️ កែសម្រួល", COLORS["primary"], self.update_staff),
            ("🗑️ លុប", COLORS["error"], self.delete_staff),
            ("🧹 សម្អាត", COLORS["warning"], self.clear_form),
            ("💾 ផ្ទុក", COLORS["accent"], self.quick_save)
        ]
        
        for text, color, command in actions:
            btn = tk.Button(action_frame, text=text, font=("Arial", 9),
                          bg=color, fg="white", relief="flat", bd=0,
                          command=command, width=8)
            btn.pack(side="left", padx=3)

    def create_enhanced_list(self, parent):
        """Create enhanced staff list with search and filters"""
        # Search and filter frame
        search_frame = tk.Frame(parent, bg=COLORS["surface"])
        search_frame.pack(fill="x", pady=(0, 10))

        # Quick search
        tk.Label(search_frame, text="🔍 ស្វែងរក:", font=("Arial", 9),
                bg=COLORS["surface"], fg=COLORS["text_primary"]).pack(side="left", padx=(0, 10))
        
        self.search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=self.search_var, font=("Arial", 9),
                              bg="#374151", fg=COLORS["text_primary"], width=25,
                              insertbackground=COLORS["text_primary"], relief="flat")
        search_entry.pack(side="left", padx=(0, 10))
        self.search_var.trace("w", lambda *args: self.search_staff())

        # Department filter
        tk.Label(search_frame, text="📊 នាយកដ្ឋាន:", font=("Arial", 9),
                bg=COLORS["surface"], fg=COLORS["text_primary"]).pack(side="left", padx=(20, 10))
        
        self.dept_filter_var = tk.StringVar()
        dept_combo = ttk.Combobox(search_frame, textvariable=self.dept_filter_var,
                                values=["ទាំងអស់", "IT", "លក់", "ទីផ្សារ", "HR", "ហិរញ្ញវត្ថុ"],
                                state="readonly", width=12)
        dept_combo.pack(side="left", padx=(0, 10))
        dept_combo.set("ទាំងអស់")
        self.dept_filter_var.trace("w", lambda *args: self.filter_staff())

        # Export buttons
        export_frame = tk.Frame(search_frame, bg=COLORS["surface"])
        export_frame.pack(side="right")
        
        tk.Button(export_frame, text="📊 Excel", font=("Arial", 8),
                 bg=COLORS["success"], fg="white", relief="flat",
                 command=self.export_to_excel).pack(side="left", padx=2)
        
        tk.Button(export_frame, text="📄 PDF", font=("Arial", 8),
                 bg=COLORS["error"], fg="white", relief="flat",
                 command=self.export_to_pdf).pack(side="left", padx=2)

        # Treeview with scrollbars
        tree_frame = tk.Frame(parent, bg=COLORS["surface"])
        tree_frame.pack(fill="both", expand=True)

        columns = ("ID", "ឈ្មោះ", "តួនាទី", "នាយកដ្ឋាន", "ប្រាក់ខែ", "ទូរស័ព្ទ", "ចូលធ្វើការ")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=15)

        # Configure columns
        col_widths = [50, 150, 120, 100, 80, 100, 100]
        for col, width in zip(columns, col_widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width)

        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        self.tree.bind("<Double-1>", self.on_tree_select)

    def browse_photo(self):
        file = filedialog.askopenfilename(
            title="ជ្រើសរើសរូបភាព",
            filetypes=[("រូបភាព", "*.jpg *.png *.jpeg *.bmp *.gif"), ("ទាំងអស់", "*.*")]
        )
        if file:
            try:
                # Resize and optimize image
                img = Image.open(file)
                img.thumbnail((400, 400))  # Resize for storage efficiency
                
                filename = f"staff_photo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
                save_to = os.path.join(self.photo_folder, filename)
                img.save(save_to, "JPEG", quality=85)
                
                self.photo_path_var.set(save_to)
                messagebox.showinfo("ជោគជ័យ", "រូបភាពត្រូវបានបន្ថែមដោយជោគជ័យ!")
            except Exception as e:
                messagebox.showerror("កំហុស", f"មិនអាចបន្ថែមរូបភាព: {e}")

    def add_staff(self):
        """Add new staff with validation"""
        if not self.name_var.get().strip():
            messagebox.showerror("កំហុស", "ឈ្មោះបុគ្គលិកត្រូវតែបំពេញ!")
            return

        # Validate salary
        try:
            salary = float(self.salary_var.get() or 0)
        except ValueError:
            messagebox.showerror("កំហុស", "ប្រាក់ខែត្រូវតែជាលេខ!")
            return

        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO staff(name, birthdate, marital_status, location, position,
                    salary, hire_date, phone, email, department, photo_path, document_path)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                self.name_var.get().strip(),
                self.birthdate_var.get().strip(),
                self.marital_var.get().strip(),
                self.location_var.get().strip(),
                self.position_var.get().strip(),
                salary,
                self.hire_date_var.get().strip(),
                self.phone_var.get().strip(),
                self.email_var.get().strip(),
                self.department_var.get().strip(),
                self.photo_path_var.get().strip(),
                self.document_path_var.get().strip()
            ))
            conn.commit()
            conn.close()
            
            self.clear_form()
            self.load_staff_data()
            self.update_stats()
            messagebox.showinfo("ជោគជ័យ", "✅ បុគ្គលិកត្រូវបានបន្ថែមដោយជោគជ័យ!")
        except Exception as e:
            messagebox.showerror("កំហុស", f"មិនអាចបន្ថែមបុគ្គលិក: {e}")

    def update_staff(self):
        if not self.selected_staff_id:
            messagebox.showerror("កំហុស", "សូមជ្រើសរើសបុគ្គលិកដើម្បីកែសម្រួល!")
            return

        try:
            salary = float(self.salary_var.get() or 0)
        except ValueError:
            messagebox.showerror("កំហុស", "ប្រាក់ខែត្រូវតែជាលេខ!")
            return

        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE staff SET
                    name=?, birthdate=?, marital_status=?, location=?, position=?,
                    salary=?, hire_date=?, phone=?, email=?, department=?,
                    photo_path=?, document_path=?, last_updated=CURRENT_TIMESTAMP
                WHERE id=?
            """, (
                self.name_var.get().strip(),
                self.birthdate_var.get().strip(),
                self.marital_var.get().strip(),
                self.location_var.get().strip(),
                self.position_var.get().strip(),
                salary,
                self.hire_date_var.get().strip(),
                self.phone_var.get().strip(),
                self.email_var.get().strip(),
                self.department_var.get().strip(),
                self.photo_path_var.get().strip(),
                self.document_path_var.get().strip(),
                self.selected_staff_id
            ))
            conn.commit()
            conn.close()
            
            self.clear_form()
            self.load_staff_data()
            self.update_stats()
            messagebox.showinfo("ជោគជ័យ", "✅ ពត៌មានបុគ្គលិកត្រូវបានកែសម្រួលដោយជោគជ័យ!")
        except Exception as e:
            messagebox.showerror("កំហុស", f"មិនអាចកែសម្រួល: {e}")

    def delete_staff(self):
        if not self.selected_staff_id:
            messagebox.showerror("កំហុស", "សូមជ្រើសរើសបុគ្គលិកដើម្បីលុប!")
            return
        
        if not messagebox.askyesno("បញ្ជាក់", "តើអ្នកពិតជាចង់លុបបុគ្គលិកនេះមែនទេ?\n\nការលុបនេះមិនអាចដកវិញបានទេ!"):
            return

        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM staff WHERE id=?", (self.selected_staff_id,))
            conn.commit()
            conn.close()
            
            self.clear_form()
            self.load_staff_data()
            self.update_stats()
            messagebox.showinfo("ជោគជ័យ", "✅ បុគ្គលិកត្រូវបានលុបដោយជោគជ័យ!")
        except Exception as e:
            messagebox.showerror("កំហុស", f"មិនអាចលុបបុគ្គលិក: {e}")

    def clear_form(self):
        """Clear all form fields"""
        self.selected_staff_id = None
        self.name_var.set("")
        self.birthdate_var.set("")
        self.marital_var.set("")
        self.location_var.set("")
        self.position_var.set("")
        self.salary_var.set("")
        self.hire_date_var.set(datetime.now().strftime("%Y-%m-%d"))
        self.phone_var.set("")
        self.email_var.set("")
        self.department_var.set("")
        self.photo_path_var.set("")
        self.document_path_var.set("")

    def load_staff_data(self):
        """Load staff data into treeview"""
        for item in self.tree.get_children():
            self.tree.delete(item)

        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, name, position, department, salary, phone, hire_date
                FROM staff ORDER BY name
            """)
            rows = cursor.fetchall()
            conn.close()

            for row in rows:
                # Format salary with commas
                formatted_row = list(row)
                formatted_row[4] = f"${row[4]:,.2f}" if row[4] else "$0.00"
                self.tree.insert("", "end", values=formatted_row)
                
            self.update_stats()
        except Exception as e:
            messagebox.showerror("កំហុស", f"មិនអាចផ្ទុកទិន្នន័យ: {e}")

    def update_stats(self):
        """Update header statistics"""
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # Total staff
            cursor.execute("SELECT COUNT(*) FROM staff")
            total = cursor.fetchone()[0]
            
            # Department count
            cursor.execute("SELECT COUNT(DISTINCT department) FROM staff WHERE department IS NOT NULL")
            dept_count = cursor.fetchone()[0]
            
            conn.close()
            
            self.total_staff_var.set(f"សរុបបុគ្គលិក: {total}")
            self.department_count_var.set(f"នាយកដ្ឋាន: {dept_count}")
            
        except Exception as e:
            print(f"Stats update error: {e}")

    def search_staff(self):
        """Search staff by name, position, or department"""
        query = self.search_var.get().lower().strip()

        for item in self.tree.get_children():
            self.tree.delete(item)

        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, name, position, department, salary, phone, hire_date
                FROM staff
                WHERE LOWER(name) LIKE ? OR LOWER(position) LIKE ? OR LOWER(department) LIKE ?
                ORDER BY name
            """, (f"%{query}%", f"%{query}%", f"%{query}%"))
            rows = cursor.fetchall()
            conn.close()

            for row in rows:
                formatted_row = list(row)
                formatted_row[4] = f"${row[4]:,.2f}" if row[4] else "$0.00"
                self.tree.insert("", "end", values=formatted_row)
        except Exception as e:
            messagebox.showerror("កំហុស", f"មិនអាចស្វែងរក: {e}")

    def filter_staff(self):
        """Filter staff by department"""
        dept = self.dept_filter_var.get()
        
        if dept == "ទាំងអស់":
            self.load_staff_data()
            return

        for item in self.tree.get_children():
            self.tree.delete(item)

        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, name, position, department, salary, phone, hire_date
                FROM staff WHERE department = ? ORDER BY name
            """, (dept,))
            rows = cursor.fetchall()
            conn.close()

            for row in rows:
                formatted_row = list(row)
                formatted_row[4] = f"${row[4]:,.2f}" if row[4] else "$0.00"
                self.tree.insert("", "end", values=formatted_row)
        except Exception as e:
            messagebox.showerror("កំហុស", f"មិនអាចច្រោះ: {e}")

    def on_tree_select(self, event):
        """Load selected staff data into form"""
        selection = self.tree.selection()
        if not selection:
            return

        staff_id = self.tree.item(selection[0])["values"][0]

        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM staff WHERE id=?", (staff_id,))
            row = cursor.fetchone()
            conn.close()

            if row:
                self.selected_staff_id = row[0]
                self.name_var.set(row[1] or "")
                self.birthdate_var.set(row[2] or "")
                self.marital_var.set(row[3] or "")
                self.location_var.set(row[4] or "")
                self.position_var.set(row[5] or "")
                self.salary_var.set(str(row[6] or ""))
                self.hire_date_var.set(row[7] or datetime.now().strftime("%Y-%m-%d"))
                self.phone_var.set(row[8] or "")
                self.email_var.set(row[9] or "")
                self.department_var.set(row[10] or "")
                self.photo_path_var.set(row[11] or "")
                self.document_path_var.set(row[12] or "")

        except Exception as e:
            messagebox.showerror("កំហុស", f"មិនអាចផ្ទុកពត៌មាន: {e}")

    def export_to_excel(self):
        """Export staff data to Excel with formatting"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("ឯកសារ Excel", "*.xlsx")],
            title="រក្សាទុករបាយការណ៍"
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "បញ្ជីបុគ្គលិក"

            # Headers with formatting
            headers = ["អត្តលេខ", "ឈ្មោះ", "ថ្ងៃកំណើត", "ស្ថានភាព", "ទីកន្លែង", 
                      "តួនាទី", "ប្រាក់ខែ", "ចូលធ្វើការ", "ទូរស័ព្ទ", "អ៊ីមែល", "នាយកដ្ឋាន"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Get data
            conn = sqlite3.connect(self.db_file)
            df = pd.read_sql_query("SELECT * FROM staff", conn)
            conn.close()

            # Write data
            for row_idx, row in df.iterrows():
                for col_idx, value in enumerate(row[:11], 1):  # First 11 columns
                    ws.cell(row=row_idx+2, column=col_idx, value=value)

            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            messagebox.showinfo("ជោគជ័យ", f"✅ របាយការណ៍ត្រូវបានរក្សាទុក:\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("កំហុស", f"មិនអាចរក្សាទុក: {e}")

    def export_to_pdf(self):
        """Simple PDF export notification"""
        messagebox.showinfo("ការអភិវឌ្ឍ", "📄 មុខងារនិងត្រូវបានអភិវឌ្ឍនៅក្នុងកំណែបន្ទាប់!")

    def quick_save(self):
        """Quick save current form"""
        if self.selected_staff_id:
            self.update_staff()
        else:
            self.add_staff()

    def preview_current_photo(self):
        """Preview current staff photo"""
        photo_path = self.photo_path_var.get()
        if photo_path and os.path.exists(photo_path):
            staff_name = self.name_var.get() or "បុគ្គលិក"
            self.show_photo_preview(photo_path, staff_name)
        else:
            messagebox.showinfo("ពត៌មាន", "រូបភាពមិនមានឬមិនត្រឹមត្រូវ!")

    def show_photo_preview(self, photo_path, staff_name):
        """Show photo preview in a dialog"""
        preview = tk.Toplevel(self.root)
        preview.title(f"🖼️ {staff_name}")
        preview.geometry("400x500")
        preview.configure(bg=COLORS["background"])
        preview.resizable(False, False)
        
        # Center window
        preview.transient(self.root)
        preview.grab_set()
        
        tk.Label(preview, text=f"🖼️ {staff_name}", font=("Arial", 14, "bold"),
                bg=COLORS["background"], fg=COLORS["text_primary"]).pack(pady=10)
        
        try:
            img = Image.open(photo_path)
            img = img.resize((300, 400), Image.Resampling.LANCZOS)
            photo_img = ImageTk.PhotoImage(img)
            
            photo_label = tk.Label(preview, image=photo_img, bg=COLORS["background"])
            photo_label.image = photo_img  # Keep reference
            photo_label.pack(pady=10)
            
        except Exception as e:
            tk.Label(preview, text="❌ មិនអាចបើករូបភាព", 
                    font=("Arial", 12), bg=COLORS["background"], 
                    fg=COLORS["error"]).pack(pady=50)
        
        tk.Button(preview, text="បិទ", font=("Arial", 10),
                 bg=COLORS["primary"], fg="white", relief="flat",
                 command=preview.destroy).pack(pady=10)

    def show_stats_view(self):
        """Show statistics view"""
        self.clear_content()
        self.current_view = "stats"
        
        stats_frame = tk.Frame(self.content_frame, bg=COLORS["surface"], padx=20, pady=20)
        stats_frame.pack(fill="both", expand=True)
        
        tk.Label(stats_frame, text="📊 របាយការណ៍ស្ថិតិបុគ្គលិក", 
                font=("Arial", 16, "bold"), bg=COLORS["surface"], 
                fg=COLORS["text_primary"]).pack(pady=10)

    def show_settings_view(self):
        """Show settings view"""
        self.clear_content()
        self.current_view = "settings"
        
        settings_frame = tk.Frame(self.content_frame, bg=COLORS["surface"], padx=20, pady=20)
        settings_frame.pack(fill="both", expand=True)
        
        tk.Label(settings_frame, text="⚙️ ការកំណត់ប្រព័ន្ធ", 
                font=("Arial", 16, "bold"), bg=COLORS["surface"], 
                fg=COLORS["text_primary"]).pack(pady=10)

    def clear_content(self):
        """Clear content frame"""
        for widget in self.content_frame.winfo_children():
            widget.destroy()

###############################################################################
# MAIN APPLICATION
###############################################################################
class MainApplication:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("🏢 ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក v2.0")
        self.root.geometry("1400x900")
        self.root.configure(bg=COLORS["background"])
        
        # Center window
        self.root.eval('tk::PlaceWindow . center')
        
        # Show welcome message
        messagebox.showinfo("ស្វាគមន៍", 
                          "🎉 ស្វាគមន៍មកកាន់ប្រព័ន្ធគ្រប់គ្រងបុគ្គលិក v2.0!\n\n"
                          "កំណែថ្មីនេះមានមុខងារដ៏ទំនើប និងងាយស្រួលប្រើប្រាស់។")
        
        # Initialize with data folder selection
        folder = get_data_folder_from_config()
        if not folder or not os.path.exists(folder):
            folder = choose_data_folder_and_save(parent=self.root)
            if not folder:
                messagebox.showerror("កំហុស", "ត្រូវការថតទិន្នន័យ!")
                self.root.destroy()
                return
        
        self.system = EnhancedStaffManagementSystem(self.root, folder)
        self.root.mainloop()

###############################################################################
# RUN ENHANCED APPLICATION
###############################################################################
if __name__ == "__main__":
    ModernLoginWindow()